"""Excel CRM generator — produces a dated spreadsheet with all analysis."""

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, MY_HANDLE


# ── colour palette ────────────────────────────────────────────────────────────
_BRAND_PURPLE = "6C3DE8"
_ACCENT_GOLD  = "F5A623"
_LIGHT_GREY   = "F4F4F6"
_MID_GREY     = "D9D9D9"
_WHITE        = "FFFFFF"
_DARK_TEXT    = "1A1A2E"
_GREEN        = "27AE60"
_BLUE_LIGHT   = "EBF5FB"
_PURPLE_LIGHT = "F3EFFE"


def _header_font(size=11, bold=True, color=_WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _cell_font(size=10, bold=False, color=_DARK_TEXT):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color=_MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)


def _wrap(ws, cell_ref, value, font=None, fill=None, align=None, border=None):
    c = ws[cell_ref]
    c.value = value
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c


# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: openpyxl.Workbook, posts: list[dict], run_date: str):
    ws = wb.active
    ws.title = "Daily Summary"

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"🎯 @{MY_HANDLE} — Instagram Content Intelligence Report"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_WHITE)
    title_cell.fill = _fill(_BRAND_PURPLE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    date_cell = ws["A2"]
    date_cell.value = f"Run date: {run_date}  |  Top {len(posts)} viral competitor videos  |  Threshold: ≥10,000 views"
    date_cell.font = _cell_font(size=9, color="666666")
    date_cell.fill = _fill(_LIGHT_GREY)
    date_cell.alignment = Alignment(horizontal="center")

    headers = [
        "Rank", "Account", "Views", "Likes", "Comments",
        "Eng. Rate %", "Posted", "URL", "Full Analysis Tab",
    ]
    col_widths = [6, 22, 14, 12, 12, 12, 18, 40, 22]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = _header_font(size=10)
        cell.fill = _fill(_BRAND_PURPLE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    row_fill = [_WHITE, _PURPLE_LIGHT]
    for rank, post in enumerate(posts, 1):
        r = rank + 3
        values = [
            f"#{rank}",
            f"@{post['username']}",
            post["views"],
            post["likes"],
            post["comments"],
            f"{post['engagement_rate']}%",
            post["posted_at"][:10],
            post["url"],
            f"Video #{rank} Detail",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = _cell_font(bold=(col == 1))
            cell.fill = _fill(row_fill[rank % 2])
            cell.border = _thin_border()
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 3, 4, 5, 6, 7, 9) else "left",
                wrap_text=True,
            )
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A4"


# ── DETAIL SHEET (one per video) ──────────────────────────────────────────────

_SPIN_KEYS = ["spin_1", "spin_2", "spin_3", "spin_4", "spin_5"]
_SPIN_LABELS = [f"Spin {i}" for i in range(1, 6)]


def _build_detail_sheet(wb: openpyxl.Workbook, post: dict, rank: int):
    sheet_name = f"Video #{rank} Detail"
    ws = wb.create_sheet(title=sheet_name)

    # ── title banner
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"#{rank} Viral Video — @{post['username']}  |  {post['views']:,} views"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=_WHITE)
    ws["A1"].fill = _fill(_BRAND_PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── metrics row
    metrics = [
        ("Views", f"{post['views']:,}"),
        ("Likes", f"{post['likes']:,}"),
        ("Comments", f"{post['comments']:,}"),
        ("Eng. Rate", f"{post['engagement_rate']}%"),
        ("Posted", post["posted_at"][:10]),
        ("URL", post["url"]),
    ]
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 50

    for col, (label, val) in enumerate(metrics, 1):
        lc = ws.cell(row=2, column=col, value=label)
        lc.font = _header_font(size=9)
        lc.fill = _fill(_ACCENT_GOLD)
        lc.alignment = Alignment(horizontal="center")

        vc = ws.cell(row=3, column=col, value=val)
        vc.font = _cell_font(size=10)
        vc.fill = _fill(_LIGHT_GREY)
        vc.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # ── caption
    _section_header(ws, 4, "Caption")
    ws.merge_cells("A5:F5")
    cap = ws["A5"]
    cap.value = post["caption"]
    cap.font = _cell_font()
    cap.fill = _fill(_WHITE)
    cap.alignment = Alignment(wrap_text=True, vertical="top")
    cap.border = _thin_border()
    ws.row_dimensions[5].height = max(60, min(len(post["caption"]) // 3, 120))

    # ── hashtags
    _section_header(ws, 6, "Hashtags")
    ws.merge_cells("A7:F7")
    ht = ws["A7"]
    ht.value = post["hashtags"]
    ht.font = _cell_font(size=9, color="555555")
    ht.fill = _fill(_LIGHT_GREY)
    ht.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 40

    # ── why it worked
    _section_header(ws, 8, "Why It Worked")
    ws.merge_cells("A9:F9")
    why = ws["A9"]
    why.value = post.get("why_it_worked", "")
    why.font = _cell_font()
    why.fill = _fill(_BLUE_LIGHT)
    why.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[9].height = 120

    # ── key patterns
    _section_header(ws, 10, "Key Patterns to Steal")
    ws.merge_cells("A11:F11")
    kp = ws["A11"]
    kp.value = post.get("key_patterns", "")
    kp.font = _cell_font()
    kp.fill = _fill(_PURPLE_LIGHT)
    kp.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[11].height = 80

    # ── spins
    _section_header(ws, 12, f"5 Ways to Spin This for @{MY_HANDLE}")

    spin_row = 13
    for i, (key, label) in enumerate(zip(_SPIN_KEYS, _SPIN_LABELS)):
        spin_fill = _fill(_LIGHT_GREY if i % 2 == 0 else _WHITE)

        header_cell = ws.cell(row=spin_row, column=1, value=f"✦ {label}")
        header_cell.font = Font(name="Calibri", bold=True, size=10, color=_BRAND_PURPLE)
        header_cell.fill = spin_fill
        ws.merge_cells(f"A{spin_row}:F{spin_row}")
        ws.row_dimensions[spin_row].height = 18

        content_row = spin_row + 1
        ws.merge_cells(f"A{content_row}:F{content_row}")
        sc = ws.cell(row=content_row, column=1, value=post.get(key, ""))
        sc.font = _cell_font()
        sc.fill = spin_fill
        sc.alignment = Alignment(wrap_text=True, vertical="top")
        sc.border = _thin_border()
        ws.row_dimensions[content_row].height = 90

        spin_row += 2

    ws.freeze_panes = "A4"


def _section_header(ws, row: int, label: str):
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value=f"  {label.upper()}")
    c.font = Font(name="Calibri", bold=True, size=10, color=_WHITE)
    c.fill = _fill(_BRAND_PURPLE)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


# ── PUBLIC ENTRY POINT ────────────────────────────────────────────────────────

def generate_crm(posts: list[dict]) -> Path:
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    filename = f"instagram_crm_{datetime.now().strftime('%Y%m%d')}.xlsx"
    out_path = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, posts, run_date)
    for rank, post in enumerate(posts, 1):
        _build_detail_sheet(wb, post, rank)

    wb.save(out_path)
    print(f"[CRM] Saved → {out_path}")
    return out_path
